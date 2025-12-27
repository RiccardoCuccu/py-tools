#!/usr/bin/env python3
"""
Social Media Follower Tracker
Continuously running script that checks Instagram and YouTube followers
on configurable schedule (daily, last day of month, or specific weekday) and exports to Excel.
Configuration stored in non-committed social_profiles.json file.
"""

import json
import os
import time
import calendar
import schedule
from datetime import datetime
from typing import Dict, Optional, Any
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Import platform scrapers
import scraper_instagram, scraper_youtube

# =====================================
# GENERAL CONFIGURATION - EDIT THESE VALUES
# =====================================

# Debug Settings
DEBUG_MODE = True  # Set to False to disable detailed debug logs

# Configuration file name (not to be committed to git)
CONFIG_FILE = "social_profiles.json"
EXCEL_FILE = "social_followers_history.xlsx"

# Get script directory for file paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, CONFIG_FILE)
EXCEL_FILE = os.path.join(SCRIPT_DIR, EXCEL_FILE)

# =====================================
# DEBUG UTILITIES
# =====================================

def debug_log(message: str) -> None:
    """Print debug message if DEBUG_MODE is enabled"""
    if DEBUG_MODE:
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[DEBUG {timestamp}] {message}")

# =====================================
# PLATFORM SCRAPERS REGISTRY
# =====================================

# Registry of available scrapers - add new platforms here
SCRAPERS = {
    "instagram": scraper_instagram.get_followers,
    "youtube": scraper_youtube.get_followers,
}

# =====================================
# FIRST RUN SETUP
# =====================================

def create_default_config() -> None:
    """Create default configuration file if it doesn't exist."""
    default_config = {
        "schedules": {
            "_comment_mode": "Options: 'daily', 'weekly', or 'monthly'",
            "mode": "monthly",
            "_comment_weekly_day": "0=Monday, 1=Tuesday, 2=Wednesday, 3=Thursday, 4=Friday, 5=Saturday, 6=Sunday (only used if mode is 'weekly')",
            "weekly_day": 0,
            "_comment_time": "Execution time in HH:MM format (24-hour)",
            "time": "23:00"
        },
        "profiles": {
            "instagram": {
                "url": "",
                "enabled": False
            },
            "youtube": {
                "url": "",
                "enabled": False
            }
        },
        "settings": {
            "retry_attempts": 3,
            "retry_delay_seconds": 5,
            "delay_between_checks_seconds": 2
        }
    }
    
    with open(CONFIG_FILE, 'w') as f:
        json.dump(default_config, f, indent=4)
    print(f"Created default config file: {CONFIG_FILE}")
    print("Please edit the file to add your profile URLs and enable tracking.")

def load_config() -> Dict[str, Any]:
    """Load configuration from JSON file."""
    if not os.path.exists(CONFIG_FILE):
        create_default_config()
        raise FileNotFoundError(f"Please configure {CONFIG_FILE} first")
    
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

# =====================================
# DATA COLLECTION
# =====================================

def collect_stats(config: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """Collect follower stats from all enabled platforms."""
    stats = {}
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    settings = config.get("settings", {})
    max_retries = settings.get("retry_attempts", 3)
    retry_delay = settings.get("retry_delay_seconds", 5)
    check_delay = settings.get("delay_between_checks_seconds", 2)
    
    for platform, profile_info in config["profiles"].items():
        if not profile_info.get("enabled", False) or not profile_info.get("url"):
            continue
        
        if platform not in SCRAPERS:
            print(f"Warning: No scraper found for platform '{platform}', skipping...")
            continue
        
        url = profile_info["url"]
        print(f"Checking {platform}: {url}")
        
        # Call the platform's scraper
        try:
            scraper_func = SCRAPERS[platform]
            result = scraper_func(url, max_retries=max_retries, retry_delay=retry_delay, debug_mode=DEBUG_MODE)
            
            stats[platform] = {
                "username": result.get("username", ""),
                "url": url,
                "followers": result.get("followers"),
                "checked_at": timestamp
            }
            
            # Delay between checks to avoid rate limiting
            if check_delay > 0 and platform != list(config["profiles"].keys())[-1]:
                time.sleep(check_delay)
                
        except Exception as e:
            print(f"Error checking {platform}: {e}")
            stats[platform] = {
                "username": "",
                "url": url,
                "followers": None,
                "checked_at": timestamp
            }
    
    return stats

# =====================================
# EXCEL REPORTING
# =====================================

def append_to_excel_history(stats: Dict[str, Dict[str, Any]]) -> None:
    """Append follower statistics to separate sheets per platform in Excel file."""
    date_str = datetime.now().strftime("%Y-%m-%d")
    time_str = datetime.now().strftime("%H:%M:%S")
    
    # Load or create workbook
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Process each platform
    for platform, data in stats.items():
        platform_name = platform.title()
        
        # Get or create sheet for this platform
        if platform_name in wb.sheetnames:
            ws = wb[platform_name]
        else:
            ws = wb.create_sheet(platform_name)
            
            # Create headers for new sheet
            headers = ["Date", "Time", "Username", "Followers", "Change"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        if ws is None:
            continue
        
        # Find last row for this sheet
        last_row = ws.max_row
        
        # Get previous value for change calculation
        previous_followers: Optional[int] = None
        if last_row > 1:
            # Get the most recent follower count
            last_followers_cell = ws.cell(row=last_row, column=4)
            if last_followers_cell and isinstance(last_followers_cell.value, int):
                previous_followers = last_followers_cell.value
        
        # Add new row
        row = last_row + 1
        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=time_str)
        ws.cell(row=row, column=3, value=data["username"])
        
        if data["followers"] is not None:
            ws.cell(row=row, column=4, value=data["followers"])
            
            # Calculate change from previous check
            if previous_followers is not None:
                change = data["followers"] - previous_followers
                change_str = f"+{change}" if change > 0 else str(change)
                change_cell = ws.cell(row=row, column=5, value=change_str)
                
                # Color code the change
                if change > 0:
                    change_cell.font = Font(color="00AA00")  # Green
                elif change < 0:
                    change_cell.font = Font(color="FF0000")  # Red
        else:
            ws.cell(row=row, column=4, value="Error")
        
        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    wb.save(EXCEL_FILE)
    print(f"History updated: {EXCEL_FILE}")

# =====================================
# SCHEDULING
# =====================================

def is_last_day_of_month() -> bool:
    """Check if today is the last day of the current month."""
    today = datetime.now()
    _, last_day = calendar.monthrange(today.year, today.month)
    return today.day == last_day

def is_target_weekday(weekday: int) -> bool:
    """Check if today matches the target weekday."""
    return datetime.now().weekday() == weekday

def should_run_job(config: Dict[str, Any]) -> bool:
    """Check if job should run based on schedule configuration."""
    sched_config = config["schedules"]
    
    if sched_config["mode"] == "daily":
        return True
    elif sched_config["mode"] == "monthly":
        return is_last_day_of_month()
    elif sched_config["mode"] == "weekly":
        return is_target_weekday(sched_config["weekly_day"])
    
    return False

def job_runner() -> None:
    """Main job: collect stats and generate report."""
    config = load_config()
    
    # Check if we should actually run
    if not should_run_job(config):
        return
    
    print(f"\n{'='*60}")
    print(f"Running follower check at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")
    
    stats = collect_stats(config)
    append_to_excel_history(stats)
    
    print(f"\n{'='*60}")
    print("Job completed successfully!")
    print(f"{'='*60}\n")

def setup_schedule(config: Dict[str, Any]) -> None:
    """Setup scheduled jobs based on configuration."""
    sched_config = config["schedules"]
    
    # Schedule to check every day at specified time
    schedule.every().day.at(sched_config["time"]).do(job_runner)
    
    if sched_config["mode"] == "daily":
        print(f"Scheduled: Daily at {sched_config['time']}")
    elif sched_config["mode"] == "monthly":
        print(f"Scheduled: Monthly (last day) at {sched_config['time']}")
    elif sched_config["mode"] == "weekly":
        weekday_name = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][sched_config["weekly_day"]]
        print(f"Scheduled: Weekly ({weekday_name}) at {sched_config['time']}")

# =====================================
# MAIN
# =====================================

def main() -> None:
    """Main function: load config, setup schedule, run forever."""
    try:
        config = load_config()
        print("\n" + "="*60)
        print("Social Media Follower Tracker Started!")
        print("="*60)
        print(f"Config loaded from: {CONFIG_FILE}")
        print(f"History file: {EXCEL_FILE}")
        print(f"Debug mode: {'ENABLED' if DEBUG_MODE else 'DISABLED'}")
        print(f"Available platforms: {', '.join(SCRAPERS.keys())}")
        
        setup_schedule(config)
        
        # Run immediately once on startup (if conditions met)
        print("\nRunning initial check...")
        job_runner()
        
        print("\nWaiting for scheduled runs...")
        print("Press Ctrl+C to stop\n")
        
        # Keep running forever
        while True:
            schedule.run_pending()
            time.sleep(60)  # Check every minute
            
    except KeyboardInterrupt:
        print("\n\nTracker stopped by user.")
    except FileNotFoundError as e:
        print(f"\n{e}")
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()